r"""Ingest Delivery Master report exports into Supabase staging tables.

Delivery Master has no API and no report scheduler, but every report it renders
lands in  <install>\Reports\<ReportType>\  as a timestamped xlsx. This script
picks up the newest export of each type and upserts it, so whoever (or whatever)
clicks Process, the data flows through without a manual upload step.

Usage:
    python dm_ingest.py                # ingest newest of each supported report
    python dm_ingest.py --list         # show what is available and how stale
    python dm_ingest.py --dry-run      # parse and report, write nothing

Credentials come from the environment, never from this file:
    SUPABASE_URL, SUPABASE_SERVICE_KEY
"""
from __future__ import annotations

import argparse
import datetime as dt
import glob
import json
import os
import sys
import urllib.error
import urllib.request

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl is required:  pip install openpyxl")

REPORTS_ROOT = r"C:\Program Files (x86)\Delivery Master\Delivery Master\Reports"

SUPABASE_URL = os.environ.get("SUPABASE_URL", "https://flvxrrwgzualfzfczlim.supabase.co")
SERVICE_KEY = os.environ.get("SUPABASE_SERVICE_KEY")


def _clean(v):
    if v is None:
        return None
    s = str(v).strip()
    return s or None


def _date(v):
    if v is None:
        return None
    if isinstance(v, (dt.datetime, dt.date)):
        return v.strftime("%Y-%m-%d")
    s = str(v).strip()
    if not s:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d %H:%M:%S"):
        try:
            return dt.datetime.strptime(s[:19], fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return None


# report type -> (target table, conflict column, row mapper)
def map_driver(r):
    call_sign = _clean(r.get("Call Sign"))
    name = _clean(r.get("Driver Name"))
    if not call_sign or not name:
        return None  # call sign is the key; rows without one cannot be joined to jobs
    return {
        "call_sign": call_sign,
        "driver_name": name,
        "driver_type": _clean(r.get("Type")),
        "postcode": _clean(r.get("Postcode")),
        "town": _clean(r.get("Town")),
    }


def map_customer(r):
    code = _clean(r.get("Account Code"))
    name = _clean(r.get("Customer Name"))
    if not code or not name:
        return None
    return {
        "account_code": code,
        "customer_name": name,
        "sales_person": _clean(r.get("Sales Person")),
        "start_date": _date(r.get("Start Date")),
        "last_trading_date": _date(r.get("Last Trading Date")),
        "nominal_code": _clean(r.get("Nominal Code")),
        "postcode": _clean(r.get("Postcode")),
        "town": _clean(r.get("Town")),
    }


REPORTS = {
    "DriverList": ("dm_drivers", "call_sign", map_driver),
    "CustomerList": ("dm_customers", "account_code", map_customer),
}


def newest_export(report_type: str):
    """Newest real export, ignoring Excel's ~$ lock files."""
    pattern = os.path.join(REPORTS_ROOT, report_type, "*")
    files = [
        f for f in glob.glob(pattern)
        if not os.path.basename(f).startswith("~$")
        and f.lower().endswith((".xlsx", ".xls"))
    ]
    return max(files, key=os.path.getmtime) if files else None


def read_rows(path: str):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    header = [str(h).strip() if h is not None else "" for h in next(rows)]
    for raw in rows:
        if raw is None or all(c is None or str(c).strip() == "" for c in raw):
            continue
        yield dict(zip(header, raw))
    wb.close()


def upsert(table: str, conflict: str, rows: list[dict]) -> int:
    if not SERVICE_KEY:
        sys.exit("SUPABASE_SERVICE_KEY is not set. Export it before running.")
    written = 0
    for i in range(0, len(rows), 500):
        batch = rows[i:i + 500]
        req = urllib.request.Request(
            f"{SUPABASE_URL}/rest/v1/{table}?on_conflict={conflict}",
            data=json.dumps(batch).encode(),
            headers={
                "apikey": SERVICE_KEY,
                "Authorization": f"Bearer {SERVICE_KEY}",
                "Content-Type": "application/json",
                "Prefer": "resolution=merge-duplicates,return=minimal",
            },
            method="POST",
        )
        try:
            with urllib.request.urlopen(req, timeout=60) as resp:
                if resp.status < 300:
                    written += len(batch)
        except urllib.error.HTTPError as e:
            sys.exit(f"{table}: HTTP {e.code} - {e.read().decode()[:300]}")
    return written


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--list", action="store_true", help="show availability and staleness only")
    ap.add_argument("--dry-run", action="store_true", help="parse but do not write")
    args = ap.parse_args()

    if args.list:
        print(f"{'REPORT':<18}{'AGE':>8}  FILE")
        for name in sorted(os.listdir(REPORTS_ROOT)):
            f = newest_export(name)
            if not f:
                continue
            age = (dt.datetime.now() - dt.datetime.fromtimestamp(os.path.getmtime(f))).days
            print(f"{name:<18}{age:>5}d  {os.path.basename(f)[:60]}")
        return

    for report_type, (table, conflict, mapper) in REPORTS.items():
        path = newest_export(report_type)
        if not path:
            print(f"{report_type:<14} no export found - run it in Delivery Master")
            continue

        stamp = dt.datetime.fromtimestamp(os.path.getmtime(path))
        age = (dt.datetime.now() - stamp).days
        src = os.path.basename(path)

        rows, skipped = [], 0
        for raw in read_rows(path):
            m = mapper(raw)
            if m is None:
                skipped += 1
                continue
            m["source_file"] = src
            rows.append(m)

        # last one wins on duplicate keys, otherwise the upsert rejects the batch
        deduped = {r[conflict]: r for r in rows}
        dupes = len(rows) - len(deduped)
        rows = list(deduped.values())

        note = f"{report_type:<14} {len(rows):>5} rows  ({age}d old"
        if skipped:
            note += f", {skipped} skipped"
        if dupes:
            note += f", {dupes} dupes collapsed"
        note += ")"

        if args.dry_run:
            print(note + "  [dry run]")
            continue

        written = upsert(table, conflict, rows)
        print(note + f"  -> {table}: {written} written")


if __name__ == "__main__":
    main()
